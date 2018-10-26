#!/usr/bin/env python3
#encode
'''
生成员工投资/推荐统计表格
:usage:
  summary.py -i requestfile -o workbookfile
'''
import json
import http.client
import sys
import getopt
import openpyxl
import time
import datetime

def getData(request):
    '''
    获取汇总数据
    :param request: 提交的汇总参数
    :return: 汇总数据
    '''
    headers = {'Content-Type': 'application/json'}
    request = json.dumps(request)
    host = 'localhost'
    port = '8080'
    method = 'POST'
    encode = 'utf-8'
    datas = [
        {'name': '公司投资总表', 'url': 'http://localhost:8090/api/rest/statistics/company-summary'},
        {'name': '员工投资&推荐汇总', 'url': 'http://localhost:8090/api/rest/statistics/company-employee-summary'},
        {'name': '员工投资明细', 'url': 'http://localhost:8090/api/rest/statistics/company-employee-invest-detail'},
        {'name': '员工推荐投资明细',
         'url': 'http://localhost:8090/api/rest/statistics/company-employee-directuser-invest-detail'}
    ]
    conn = http.client.HTTPConnection(host, port)

    for data in datas:
        conn.request(method, data['url'], request.encode(encode), headers)
        response = conn.getresponse()
        result = json.loads(response.read().decode(encode))
        data.update({'data': result})

    return datas


def generateTableFile(data, outputFile):
    '''
    生成工作薄文件
    :param data: 汇总数据
    :param outputFile: 工作薄文件名
    :return:
    '''
    workbook = openpyxl.workbook.Workbook()
    generateCompanySummaryTable(data[0], workbook)
    generateCompanyEmployeeSummaryTable(data[1], workbook)
    generateCompanyEmployeeInvestTable(data[2], workbook)
    generateCompanyEmployeeDirectuserInvestTable(data[3], workbook)
    workbook.save(outputFile)
    print('\n')
    print(data)


def generateCompanySummaryTable(data, workbook):
    '''
    生成公司汇总工作表
    :param data: 公司汇总数据
    :param workbook: 汇总工作薄
    :return:
    '''
    sheet = workbook.active
    sheet.title = data['name']

    ds = data['data']['data']

    sheet.cell(1, 1).value = '时间范围'
    sheet.cell(1, 2).value = ds['startTime'][0:10]
    sheet.cell(1, 3).value = ds['endTime'][0:10]

    sheet.cell(3, 1).value = '公司名称'
    sheet.cell(3, 2).value = '员工人数'
    sheet.cell(3, 3).value = '投资人数'

    i = 4

    for row in ds['companyList']:
        sheet.cell(i, 1).value = row['companyName']
        sheet.cell(i, 2).value = row['total']
        sheet.cell(i, 3).value = row['investNumber']
        i += 1

    render(4, i, sheet)


def render(f, i, sheet):
    '''
    渲染工作表批指定区域
    :param f: 结束列
    :param i: 结束行
    :param sheet:
    :return:
    '''
    ft = openpyxl.styles.Font(name='微软雅黑', size=10)
    tft = openpyxl.styles.Font(name='微软雅黑', size=11)
    ali = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    tfill = openpyxl.styles.PatternFill(start_color='DDDDDDDD', end_color='DDDDDDDD', fill_type='solid')
    border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin', color='FF000000'),
                                    right=openpyxl.styles.Side(style='thin', color='FF000000'),
                                    top=openpyxl.styles.Side(style='thin', color='FF000000'),
                                    bottom=openpyxl.styles.Side(style='thin', color='FF000000'),
                                    diagonal=openpyxl.styles.Side(style='thin', color='FF000000'),
                                    diagonal_direction=0,
                                    outline=openpyxl.styles.Side(style='thin', color='FF000000'),
                                    vertical=openpyxl.styles.Side(style='thin', color='FF000000'),
                                    horizontal=openpyxl.styles.Side(style='thin', color='FF000000'))
    for r in range(1, i):
        for c in range(1, f):
            if (r == 3):
                sheet.cell(r, c).font = tft
                sheet.cell(r, c).fill = tfill
            else:
                sheet.cell(r, c).font = ft
            sheet.cell(r, c).alignment = ali
            if (r != 1 and r != 2):
                sheet.cell(r, c).border = border
            col = sheet.column_dimensions['abcdefghijklmn'[c - 1]]
            col.width = 24
        row = sheet.row_dimensions[r]
        row.height = 16

def generateCompanyEmployeeSummaryTable(data, workbook):
    '''
    生成员工投资/推荐汇总工作表
    :param data: 员工投资/推荐汇总数据
    :param workbook: 汇总工作薄
    :return:
    '''
    sheet = workbook.create_sheet(data['name'])

    ds = data['data']['data']

    sheet.cell(1, 1).value = '时间范围'
    sheet.cell(1, 2).value = ds['startTime'][0:10]
    sheet.cell(1, 3).value = ds['endTime'][0:10]

    sheet.cell(3, 1).value = '公司名称'
    sheet.cell(3, 2).value = '员工姓名'
    sheet.cell(3, 3).value = '手机号码'
    sheet.cell(3, 4).value = '邀请码'
    sheet.cell(3, 5).value = '投资总额'
    sheet.cell(3, 6).value = '被邀人总数'
    sheet.cell(3, 7).value = '被邀人投资总额'

    i = 4

    for company in ds['companyList']:
        employeesNumber = len(company['employeeSummaryList'])
        if employeesNumber > 1: sheet.merge_cells('a' + str(i) + ':' + 'a' + str(i + employeesNumber - 1))
        sheet.cell(i, 1).value = company['companyName']
        for employee in company['employeeSummaryList']:
            sheet.cell(i, 2).value = employee['name']
            sheet.cell(i, 3).value = employee['mobile']
            sheet.cell(i, 4).value = employee['directCode']
            sheet.cell(i, 5).value = employee['investTotal']
            sheet.cell(i, 6).value = employee['directuserTotal']
            sheet.cell(i, 7).value = employee['directuserInvestTotal']
            i += 1

    render(8, i, sheet)

def generateCompanyEmployeeInvestTable(data, workbook):
    '''
    生成员工投资明细工作表
    :param data: 员工投资明细数据
    :param workbook: 汇总工作薄
    :return:
    '''
    sheet = workbook.create_sheet(data['name'])

    ds = data['data']['data']

    sheet.cell(1, 1).value = '时间范围'
    sheet.cell(1, 2).value = ds['startTime'][0:10]
    sheet.cell(1, 3).value = ds['endTime'][0:10]

    sheet.cell(3, 1).value = '公司名称'
    sheet.cell(3, 2).value = '员工姓名'
    sheet.cell(3, 3).value = '手机号码'
    sheet.cell(3, 4).value = '投资金额'
    sheet.cell(3, 5).value = '投资项目'
    sheet.cell(3, 6).value = '投资时间'

    i = 4

    for company in ds['companyList']:
        k = i
        sheet.cell(i, 1).value = company['companyName']
        for employee in company['investUserList']:
            j = i
            sheet.cell(i, 2).value = employee['name']
            sheet.cell(i, 3).value = employee['mobile']
            if employee['investDetailList']:
                for invest in employee['investDetailList']:
                    sheet.cell(i, 4).value = invest['invest']
                    sheet.cell(i, 5).value = invest['title']
                    sheet.cell(i, 6).value = invest['investDate'][0:10]
                    i += 1
                sheet.merge_cells('b' + str(j) + ':' + 'b' + str(i - 1))
                sheet.merge_cells('c' + str(j) + ':' + 'c' + str(i - 1))
            else:
                i += 1
        sheet.merge_cells('a' + str(k) + ':' + 'a' + str(i - 1))

    render(7, i, sheet)

def generateCompanyEmployeeDirectuserInvestTable(data, workbook):
    '''
    生成被邀人投资详情工作表
    :param data: 被邀人投资详情数据
    :param workbook: 汇总工作薄
    :return:
    '''
    sheet = workbook.create_sheet(data['name'])

    ds = data['data']['data']

    sheet.cell(1, 1).value = '时间范围'
    sheet.cell(1, 2).value = ds['startTime'][0:10]
    sheet.cell(1, 3).value = ds['endTime'][0:10]

    sheet.cell(3, 1).value = '公司名称'
    sheet.cell(3, 2).value = '员工姓名'
    sheet.cell(3, 3).value = '手机号码'
    sheet.cell(3, 4).value = '被邀人'
    sheet.cell(3, 5).value = '被邀人手机'
    sheet.cell(3, 6).value = '被邀人投资金额'
    sheet.cell(3, 7).value = '被邀人投资项目'
    sheet.cell(3, 8).value = '被邀人投资时间'

    i = 4

    for company in ds['companyList']:
        k = i
        sheet.cell(i, 1).value = company['companyName']
        for employee in company['investUserList']:
            j = i
            sheet.cell(i, 2).value = employee['name']
            sheet.cell(i, 3).value = employee['mobile']
            if employee['investUserList']:
                for directuser in employee['investUserList']:
                    m = i
                    sheet.cell(i, 4).value = directuser['name']
                    sheet.cell(i, 5).value = directuser['mobile']
                    if directuser['investDetailList']:

                        for invest in directuser['investDetailList']:
                            sheet.cell(i, 6).value = invest['invest']
                            sheet.cell(i, 7).value = invest['title']
                            sheet.cell(i, 8).value = invest['investDate'][0:10]
                            i += 1
                        sheet.merge_cells('d' + str(m) + ':' + 'd' + str(i - 1))
                        sheet.merge_cells('e' + str(m) + ':' + 'e' + str(i - 1))
                    else:
                        i += 1
                sheet.merge_cells('b' + str(j) + ':' + 'b' + str(i - 1))
                sheet.merge_cells('c' + str(j) + ':' + 'c' + str(i - 1))
            else:
                i += 1
        sheet.merge_cells('a' + str(k) + ':' + 'a' + str(i - 1))

    render(9, i, sheet)

def loadRequest(file):
    '''
    载入参数文件，并将其转换为要提交给数据提供子系统的request对象
    :param file: 参数文件
    :return: request对象
    '''
    workbook = openpyxl.load_workbook(file)

    sheetnames = workbook.sheetnames

    sheet = workbook[sheetnames[0]]
    print(sheet.title)

    startTime = datetime2timestamp(sheet.cell(1, 2).value)
    print(startTime)

    endTime = datetime2timestamp(sheet.cell(1, 3).value)
    print(endTime)

    mergecells = sheet.merged_cells
    print(mergecells)

    companyList = []

    readRow = 4

    for mergecell in mergecells:
        company = {}
        if mergecell.min_row < 3:
            continue
        if mergecell.min_row > readRow:
            for row in range(readRow, mergecell.min_row):
                com = {}
                com.update({"companyName": sheet.cell(row, 1).value})
                investUserList = [{"name": sheet.cell(row, 2).value, "mobile": sheet.cell(row, 3).value}]
                com.update({'investUserList': investUserList})
                companyList.append(com)
        readRow = mergecell.max_row + 1

        company.update({"companyName": sheet.cell(mergecell.min_row, 1).value})
        investUserList = []
        for row in range(mergecell.min_row, mergecell.max_row + 1):
            investUser = {}
            investUser.update({
                "name": sheet.cell(row, 2).value,
                "mobile": sheet.cell(row, 3).value
            })
            print(investUser)
            investUserList.append(investUser)
        company.update({"investUserList": investUserList})
        print(company)

        companyList.append(company)

    print(companyList)
    request = {"startTime": startTime, "endTime": endTime, "companyList": companyList}
    return request

def datetime2timestamp(datetime):
    '''
    将日期轮换为时间戳（毫秒）
    :param datetime: 日期，格式为 yyyy-MM-dd HH:mm:ss
    :return: 时间戳（毫秒）
    '''
    return int(time.mktime(time.strptime(str(datetime), "%Y-%m-%d %H:%M:%S"))) * 1000

def parseArgs():
    '''
    解析命令行参数
    :return: inputFile: 参数文件名；outputFile：输出文件名
    '''
    opts, args = getopt.getopt(sys.argv[1:], 'i:o:')
    inputFile = ''
    outputFile = ''
    for op, value in opts:
        if op == '-i':
            inputFile = value
        elif op == '-o':
            outputFile = value
    if inputFile == '' or outputFile == '':
        return False
    else:
        return inputFile, outputFile

inputFile, outputFile = parseArgs()
print(inputFile, outputFile)
request = loadRequest(inputFile)
data = getData(request)
generateTableFile(data, outputFile)
